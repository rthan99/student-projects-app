import sys
from pathlib import Path
from importlib.machinery import SourceFileLoader
import sqlite3


def pick(d, key):
    for k in [key, key.lower(), key.replace(" ", "_"), key.title(), key.upper()]:
        if k in d and str(d[k]).strip() != "":
            return d[k]
    return None


def import_xlsx(xlsx_path: Path, create_project_fn) -> None:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise SystemExit("Excel support requires 'openpyxl'. Install with: python3 -m pip install openpyxl") from exc

    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    # Use row 2 as headers per request
    header_cells = next(ws.iter_rows(min_row=2, max_row=2))
    headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
    header_lower = [h.lower() for h in headers]

    created = 0
    skipped = 0
    # Data starts from row 3
    for r in ws.iter_rows(min_row=3, values_only=True):
        item = {}
        for idx, val in enumerate(r):
            key = header_lower[idx] if idx < len(header_lower) else f"col{idx}"
            item[key] = val if val is not None else ""
        try:
            # Map NEMO spreadsheet columns
            title = item.get("project title") or item.get("title")
            description = item.get("short description   (note with further details)") or item.get("description")
            category = item.get("physics themes") or item.get("category")
            date_val = item.get("date")
            year = None
            if date_val:
                try:
                    s = str(date_val).strip()
                    if len(s) == 4 and s.isdigit():
                        year = int(s)
                    elif "-" in s:
                        end = s.split("-")[-1]
                        yy = int(''.join(filter(str.isdigit, end))[-2:])
                        year = 2000 + yy if yy < 70 else 1900 + yy
                except Exception:
                    year = None

            if not title:
                skipped += 1
                continue

            tags = []
            if item.get("interaction types"):
                tags.append(str(item.get("interaction types")).strip())
            if item.get("status"):
                tags.append(str(item.get("status")).strip())

            create_project_fn(
                str(title).strip(),
                "NEMO x Delft",
                str(category).strip() if category else None,
                [t for t in tags if t],
                str(description).strip() if description else None,
                year,
            )
            created += 1
        except Exception:
            skipped += 1
    print(f"Created: {created}, Skipped: {skipped}")


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python3 import_projects.py /absolute/path/to/file.xlsx")
        sys.exit(1)
    path = Path(sys.argv[1]).expanduser().resolve()
    if not path.exists():
        print(f"File not found: {path}")
        sys.exit(1)

    # Load db and models without importing app.__init__ (which requires Flask)
    project_root = Path(__file__).resolve().parent
    db_mod = SourceFileLoader("db", str(project_root / "app" / "db.py")).load_module()

    def create_project_impl(title: str, student_name: str, category=None, tags=None, description=None, year=None) -> int:
        tags_str = ",".join(tags) if isinstance(tags, list) else (tags if tags else None)
        with db_mod.get_connection() as connection:  # type: ignore[attr-defined]
            cursor = connection.cursor()
            cursor.execute(
                (
                    "INSERT INTO projects (title, student_name, category, tags, description, year) "
                    "VALUES (?, ?, ?, ?, ?, ?)"
                ),
                (title, student_name, category, tags_str, description, year),
            )
            return cursor.lastrowid

    # Ensure DB exists
    db_mod.initialize_database()  # type: ignore[attr-defined]

    if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        import_xlsx(path, create_project_impl)
    else:
        print("Unsupported file type. Please provide an .xlsx file.")
        sys.exit(1)


if __name__ == "__main__":
    main()


