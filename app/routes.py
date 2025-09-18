import os
from pathlib import Path
from flask import Blueprint, current_app, jsonify, request, send_from_directory
from werkzeug.utils import secure_filename
from .models import create_project, list_projects, get_project, add_media, delete_project, update_project, get_media, delete_media
import csv
from io import StringIO, BytesIO


bp = Blueprint("api", __name__)


@bp.route("/api/projects", methods=["GET"])
def api_list_projects():
    projects = list_projects(
        search=request.args.get("q"),
        category=request.args.get("category"),
        year=int(request.args["year"]) if request.args.get("year") else None,
        rating=int(request.args["rating"]) if request.args.get("rating") else None,
        sort_by=request.args.get("sort", "created_at"),
        order=request.args.get("order", "desc"),
    )
    return jsonify(projects)


@bp.route("/api/projects", methods=["POST"])
def api_create_project():
    data = request.form if request.form else request.json
    if data is None:
        return jsonify({"error": "No data provided"}), 400
    title = data.get("title")
    student_name = data.get("student_name")
    if not title or not student_name:
        return jsonify({"error": "title and student_name are required"}), 400
    categories = data.get("categories")
    if isinstance(categories, str):
        categories = [cat.strip() for cat in categories.split(",") if cat.strip()]
    elif not isinstance(categories, list):
        categories = []
    
    tags = data.get("tags")
    if isinstance(tags, str):
        tags = [t.strip() for t in tags.split(",") if t.strip()]
    description = data.get("description")
    year = int(data["year"]) if data and data.get("year") else None
    video_url = data.get("video_url")
    rating = int(data["rating"]) if data and data.get("rating") else 0
    project_id = create_project(title, student_name, categories, tags, description, year, video_url, rating)

    # If multipart form with files, handle optional image and video uploads here
    if request.files:
        # Image upload
        image_file = request.files.get("image")
        if image_file and image_file.filename:
            if not allowed_file(image_file.filename, is_image=True):
                return jsonify({"error": "Unsupported image type"}), 400
            img_filename = secure_filename(image_file.filename)
            img_save_name = f"{project_id}_img_{img_filename}"
            image_dir = Path(current_app.config["UPLOAD_FOLDER_IMAGES"]) 
            image_dir.mkdir(parents=True, exist_ok=True)
            image_file.save(image_dir / img_save_name)
            add_media(project_id, "image", img_save_name, image_file.filename)

        # Video upload
        video_file = request.files.get("video")
        if video_file and video_file.filename:
            if not allowed_file(video_file.filename, is_image=False):
                return jsonify({"error": "Unsupported video type"}), 400
            vid_filename = secure_filename(video_file.filename)
            vid_save_name = f"{project_id}_vid_{vid_filename}"
            video_dir = Path(current_app.config["UPLOAD_FOLDER_VIDEOS"]) 
            video_dir.mkdir(parents=True, exist_ok=True)
            video_file.save(video_dir / vid_save_name)
            add_media(project_id, "video", vid_save_name, video_file.filename)

    return jsonify({"id": project_id}), 201


@bp.route("/api/projects/<int:project_id>", methods=["GET"])
def api_get_project(project_id: int):
    project = get_project(project_id)
    if project is None:
        return jsonify({"error": "Not found"}), 404
    return jsonify(project)


@bp.route("/api/projects/<int:project_id>", methods=["DELETE"])
def api_delete_project(project_id: int):
    delete_project(project_id)
    return ("", 204)
@bp.route("/api/projects/<int:project_id>", methods=["PUT"])
def api_update_project(project_id: int):
    data = request.json
    if data is None:
        return jsonify({"error": "No data provided"}), 400
    title = data.get("title") if "title" in data else None
    student_name = data.get("student_name") if "student_name" in data else None
    categories = data.get("categories") if "categories" in data else None
    if categories and isinstance(categories, str):
        categories = [cat.strip() for cat in categories.split(",") if cat.strip()]
    elif categories and not isinstance(categories, list):
        categories = []
    
    tags = data.get("tags") if "tags" in data else None
    if isinstance(tags, str):
        tags = [t.strip() for t in tags.split(",") if t.strip()]
    description = data.get("description") if "description" in data else None
    year = data.get("year") if "year" in data else None
    if isinstance(year, str) and year.isdigit():
        year = int(year)
    video_url = data.get("video_url") if "video_url" in data else None
    rating = data.get("rating") if "rating" in data else None
    if isinstance(rating, str) and rating.isdigit():
        rating = int(rating)
    update_project(project_id, title, student_name, categories, tags, description, year, video_url, rating)
    return ("", 204)



ALLOWED_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}
ALLOWED_VIDEO_EXTENSIONS = {"mp4", "mov", "webm", "ogg", "mkv"}


def allowed_file(filename: str, is_image: bool) -> bool:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext in (ALLOWED_IMAGE_EXTENSIONS if is_image else ALLOWED_VIDEO_EXTENSIONS)


@bp.route("/api/projects/<int:project_id>/upload/image", methods=["POST"])
def api_upload_image(project_id: int):
    if "file" not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400
    if not allowed_file(file.filename, is_image=True):
        return jsonify({"error": "Unsupported image type"}), 400
    filename = secure_filename(file.filename)
    save_name = f"{project_id}_img_{filename}"
    image_dir = Path(current_app.config["UPLOAD_FOLDER_IMAGES"]) 
    image_dir.mkdir(parents=True, exist_ok=True)
    save_path = image_dir / save_name
    file.save(save_path)
    add_media(project_id, "image", save_name, file.filename)
    return jsonify({"filename": save_name})


@bp.route("/api/projects/<int:project_id>/upload/video", methods=["POST"])
def api_upload_video(project_id: int):
    if "file" not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400
    if not allowed_file(file.filename, is_image=False):
        return jsonify({"error": "Unsupported video type"}), 400
    filename = secure_filename(file.filename)
    save_name = f"{project_id}_vid_{filename}"
    video_dir = Path(current_app.config["UPLOAD_FOLDER_VIDEOS"]) 
    video_dir.mkdir(parents=True, exist_ok=True)
    save_path = video_dir / save_name
    file.save(save_path)
    add_media(project_id, "video", save_name, file.filename)
    return jsonify({"filename": save_name})


@bp.route("/uploads/images/<path:filename>")
def serve_image(filename: str):
    return send_from_directory(current_app.config["UPLOAD_FOLDER_IMAGES"], filename)


@bp.route("/uploads/videos/<path:filename>")
def serve_video(filename: str):
    return send_from_directory(current_app.config["UPLOAD_FOLDER_VIDEOS"], filename)


@bp.route("/api/media/<int:media_id>", methods=["DELETE"])
def api_delete_media(media_id: int):
    media = get_media(media_id)
    if not media:
        return jsonify({"error": "Not found"}), 404
    # delete file from disk
    try:
        base = current_app.config["UPLOAD_FOLDER_IMAGES"] if media["media_type"] == "image" else current_app.config["UPLOAD_FOLDER_VIDEOS"]
        path = Path(base) / media["filename"]
        if path.exists():
            path.unlink()
    except Exception:
        pass
    delete_media(media_id)
    return ("", 204)

@bp.route("/api/projects/import", methods=["POST"])
def api_import_projects():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    upload = request.files["file"]
    if upload.filename == "":
        return jsonify({"error": "Empty filename"}), 400

    filename = secure_filename(upload.filename)
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    rows = []
    try:
        if ext == "csv":
            text = upload.stream.read().decode("utf-8", errors="ignore")
            reader = csv.DictReader(StringIO(text))
            rows = list(reader)
        elif ext in {"xlsx", "xlsm", "xltx", "xltm"}:
            try:
                from openpyxl import load_workbook  # type: ignore
            except Exception:
                return jsonify({"error": "Excel support requires 'openpyxl' to be installed."}), 400
            wb = load_workbook(filename=BytesIO(upload.read()), read_only=True, data_only=True)
            ws = wb.active
            headers = []
            # Use row 2 as headers per request
            for cell in next(ws.iter_rows(min_row=2, max_row=2)):
                headers.append(str(cell.value).strip() if cell.value is not None else "")
            header_lower = [h.lower() for h in headers]
            # Data starts from row 3
            for r in ws.iter_rows(min_row=3, values_only=True):
                row_map = {}
                for idx, val in enumerate(r):
                    key = header_lower[idx] if idx < len(header_lower) else f"col{idx}"
                    row_map[key] = val if val is not None else ""
                # Map known NEMO columns to our schema
                def get(key: str):
                    return row_map.get(key, None)

                title = get("project title") or get("title")
                description = get("short description   (note with further details)") or get("description")
                category = get("physics themes") or get("category")
                date_val = get("date")
                year = None
                if date_val:
                    try:
                        s = str(date_val).strip()
                        if len(s) == 4 and s.isdigit():
                            year = int(s)
                        elif "-" in s:
                            # e.g., 23-24 -> 2024
                            end = s.split("-")[-1]
                            yy = int(''.join(filter(str.isdigit, end))[-2:])
                            year = 2000 + yy if yy < 70 else 1900 + yy
                    except Exception:
                        year = None

                tags = []
                if row_map.get("interaction types"):
                    tags.append(str(row_map.get("interaction types")).strip())
                if row_map.get("status"):
                    tags.append(str(row_map.get("status")).strip())

                rows.append({
                    "title": title,
                    "student_name": "NEMO x Delft",
                    "category": category or None,
                    "tags": ",".join([t for t in tags if t]) if tags else None,
                    "description": description or None,
                    "year": year,
                })
        else:
            return jsonify({"error": "Unsupported file type. Use CSV or XLSX."}), 400
    except Exception as e:
        return jsonify({"error": f"Failed to parse file: {e}"}), 400

    def pick(d, key):
        for k in [key, key.lower(), key.replace(" ", "_"), key.title(), key.upper()]:
            if k in d and str(d[k]).strip() != "":
                return d[k]
        return None

    created = 0
    errors = []
    for idx, raw in enumerate(rows, start=2):
        try:
            title = pick(raw, "title")
            student_name = pick(raw, "student_name") or pick(raw, "student") or pick(raw, "student name")
            if not title or not student_name:
                raise ValueError("Missing required 'title' or 'student_name'")
            category = pick(raw, "category")
            tags_raw = pick(raw, "tags")
            tags = [t.strip() for t in str(tags_raw).split(",") if t.strip()] if tags_raw else None
            description = pick(raw, "description") or pick(raw, "desc")
            year_val = pick(raw, "year")
            try:
                year = int(year_val) if year_val not in (None, "") else None
            except Exception:
                year = None
            create_project(str(title).strip(), str(student_name).strip(), str(category).strip() if category else None, tags, str(description).strip() if description else None, year)
            created += 1
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})

    return jsonify({"created": created, "errors": errors})

